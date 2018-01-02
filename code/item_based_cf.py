import os
import openpyxl

def IsScore(sheet,r,c):
    if sheet.cell(row = r + 1, column=c).value != 'NT' \
            and sheet.cell(row = r + 1, column=c).value != 'I' \
            and sheet.cell(row = r + 1, column=c).value != 'XX' \
            and sheet.cell(row = r + 1, column=c).value != None:
        return True
    else:
        return False

def scr_sub(sheet,u,i):
    return sheet.cell(row = u+1,column = i).value

print("Collaborative Filtering Item-Based Algorithm for Grade Prediction")
wb= openpyxl.load_workbook('data/student.xlsx')
# print( type(wb))

sheetname = wb.get_sheet_names()
# print(sheetname[0])
print("Student's score matrix(roll no in first column)\n")

sheet=wb.get_sheet_by_name(sheetname[1])
# print(sheet)
for i in range(2,sheet.max_row+1,1):
    for j in range(1,sheet.max_column+1,1):
        if sheet.cell(row=i,column=j).value != None:
            if j == 1:
                if int(sheet.cell(row=i,column=j).value/10)==0:
                    print(sheet.cell(row=i, column=j).value, end="  | ")
                else:
                    print(sheet.cell(row=i,column=j).value,end=" |  ")
            else:
                print(sheet.cell(row=i,column=j).value,end="  ")
            if sheet.cell(row=i, column=j).value != 'I' and sheet.cell(row=i, column=j).value != 'XX' and sheet.cell(row=i, column=j).value != 'NT':
                if int(sheet.cell(row=i,column=j).value/10) == 0:
                    print(end=" ")
            if sheet.cell(row=i,column=j).value=='I':
                print(end=" ")
        else:
            print("-",end="   ")
    print()
studid = int(input("Enter roll no. : "))
# print(studid);
coursesList = []
for i in range (2,sheet.max_column+1):
    if (i-2)%9 == 0:
        print()
    print(sheet.cell(row=1,column=i).value,end=", ")
    coursesList.append(sheet.cell(row=1,column=i).value)
print("\n")
while True:
    try:
        c = input("\nEnter course: ")
        print(c)
        course = coursesList.index(c)+1
        break
    except ValueError:
        print("Please choose a course from given list")
###---------item based-------
#similarity b/w u,courses
# avg grade of every student
def getAvgGrades(sheet,student):    # calculate avg grades for student
    sum = 0
    cnt = 0
    for i in range (2,sheet.max_column+1):
        if IsScore(sheet,student,i):
            sum = sum + int(scr_sub(sheet,student,i))
            cnt = cnt + 1
    if cnt != 0:
        return round(float(sum/cnt),3)
    else:
        return 0


def similarity(c1, c2, sheet):
    sum1 = 0
    sum2 = 0
    sum3 = 0
    # avgV = getAvgGrades(sheet, v)
    # print(" avg ",avgU,avgV,v)
    for i in range(1,sheet.max_row):
        if IsScore(sheet,i,c1) and IsScore(sheet,i,c2):
            avg = getAvgGrades(sheet,i)
            # print(i,avg)
            sum1 = sum1 + (scr_sub(sheet,i,c1) - avg) * (scr_sub(sheet,i,c2) - avg)
            sum2 = sum2 + (scr_sub(sheet,i,c1) - avg) * (scr_sub(sheet,i,c1) - avg)
            sum3 = sum3 + (scr_sub(sheet,i,c2) - avg) * (scr_sub(sheet,i,c2) - avg)
    if sum2!=0 and sum3!=0:
        return round(float(sum1/((sum2*sum3)**(1/2))),3)
    else:
        return 0


similar_item_val = []
simi_item = []
avgU = getAvgGrades(sheet, studid)
for i in range(2,sheet.max_column+1):
        if i!=course:
            sim = similarity(i,course,sheet)
            similar_item_val.append([course,i,sim])
        if sim > 0:
            simi_item.append([course,i,sim])

# print(similar_item_val)
# print(simi_item)


#predicting grade
sum1 = 0
sum2 = 0
for i in range(0,len(simi_item)):
    if IsScore(sheet,studid,int(simi_item[i][1])):
        sum1 = sum1 + (simi_item[i][2] * scr_sub(sheet,studid,simi_item[i][1]))
        sum2 = sum2 + simi_item[i][2]
if sum2 != 0:
    print("Predicted Grade for student "+str(studid)+" in subject "+str(c)+" : "+str(round(float(sum1/sum2),3)))