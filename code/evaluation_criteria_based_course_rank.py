import os
import openpyxl


print("Ranked List of Course based on evaluation criteria preferences of student")
wb= openpyxl.load_workbook('data/Data_1.xlsx')


sheetname = wb.get_sheet_names()
# print(sheetname[1])

# rsheet=wb.get_sheet_by_name(sheetname[0])
wsheet=wb.get_sheet_by_name(sheetname[1])
# print("Enter Sudent ID")
studid = int(input("Enter Student ID : "))
list =[]
course = []
for i in range(2, 27):
    # print(wsheet.cell(row=studid+1,column=i).value, wsheet.cell(row=studid+1,column=28).value, studid+1,i)
    if wsheet.cell(row=studid+1,column=i).value - wsheet.cell(row=studid+1,column=28).value >0:
        list.append(round((wsheet.cell(row=studid+1,column=i).value - wsheet.cell(row=studid+1,column=28).value),3))
        course.append(i)
# print(list)
# print(course)

for i in range(0,len(list)):
    for j in range(0,len(list)-1):
        if list[j]<list[j+1]:
            tmp = list[j]
            list[j] = list[j+1]
            list[j+1] = tmp
            tmp = course[j]
            course[j] = course[j+1]
            course[j+1] = tmp

# print(list)
# print(course)
print("\nList of courses\n")
for i in range(0,len(course)):
    print(i+1,wsheet.cell(row=1,column=course[i]).value)
# print("done")
