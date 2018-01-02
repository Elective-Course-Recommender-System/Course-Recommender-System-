import os
import openpyxl


#print("Collaborative Filtering Item-Based Algorithm for Grade Prediction")
wb= openpyxl.load_workbook('data/Data_1.xlsx')


sheetname = wb.get_sheet_names()
print(sheetname[0])

rsheet=wb.get_sheet_by_name(sheetname[0])
wsheet=wb.get_sheet_by_name(sheetname[1])

def multiply(rsheet,wsheet,s,c,avgsum): # as per sheet
    sum = 0
    for j in range(2,7):
        sum = sum + rsheet.cell(row = s+1,column= j ).value * rsheet.cell(row=c+1,column = j+7).value
        # print(s+1,c+1,rsheet.cell(row = s+1,column= j ).value,rsheet.cell(row=c+1,column = j+7).value)
    wsheet.cell(row=s+1,column=c+1).value = sum
    # avgsum = sum+avgsum
# print(rsheet.max_column,rsheet.max_row)
for i in range (1,rsheet.max_row):
    avgsum = 0
    for j in range(1,26):
        multiply(rsheet,wsheet,i,j,avgsum)
        # print('out', avgsum)


wb.save('data/Data_1.xlsx')
#avrage
for i in range(2,wsheet.max_row+1):
    sum = 0
    for j in range(2,27):
        if(wsheet.cell(row=i,column=j).value==None):
            print(i,j)
        else:
            print(wsheet.cell(row=i,column=j).value, i,j)
            sum = sum + wsheet.cell(row=i,column=j).value
    wsheet.cell(row=i, column=28).value = round(float(sum/25),3)
wb.save('data/Data_1.xlsx')

print("done")
