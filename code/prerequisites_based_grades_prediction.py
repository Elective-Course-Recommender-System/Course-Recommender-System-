import os
import openpyxl

wb = openpyxl.load_workbook('data/prerequsite.xlsx')
wb2 = openpyxl.load_workbook('data/student.xlsx')
sheetname = wb.get_sheet_names()
sheetnamestudent = wb2.get_sheet_names()

grade = wb2.get_sheet_by_name(sheetnamestudent[0])
cpi = wb.get_sheet_by_name(sheetname[2])
weight = wb.get_sheet_by_name(sheetname[0])
wsheet = wb.get_sheet_by_name(sheetname[1])

course =[]
for i in range (2,grade.max_column+1):
    course.append(grade.cell(row=1,column=i).value)


for i in range(1,54):
    for j in range(2,7):
        sum = 0
        for k in range(2,weight.max_column+1):
            if weight.cell(row=j,column=k).value>0:
                if grade.cell(row=i+1,column=k).value!= None and grade.cell(row=i+1,column=k).value!= 'XX' and grade.cell(row=i + 1, column=k).value != 'I' and grade.cell(row=i+1,column=k).value!= 'NT':
                    sum = sum + (grade.cell(row=i+1,column=k).value - cpi.cell(row=i+1,column=2).value) * weight.cell(row = j,column = k).value

        wsheet.cell(row=i+1,column=j).value = cpi.cell(row=i + 1, column=2).value + sum
wb.save('data/prerequsite.xlsx')
print('Grades are calculated and stored in "data/prerequsite.xlsx" file')
